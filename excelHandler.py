import json
from sqlite3 import IntegrityError

import openpyxl as xl

def handlingFields(list: list, *args):
    """
    if each value of *args is not equal 'none', and,
    list dosen't contain the list of args, then
    append to the list the list of args.
    """
    # for item in args:
    #     if item == 'none':
    #         return

    if args not in list:
        list.append(args)

def proccess_workbook(filename, sheetname):
    wb = xl.load_workbook(filename)
    sheet = wb[sheetname]

    list_items = []
    for row in range(2, sheet.max_row + 1):
        item = {}
        for col in range(1, sheet.max_column):
            # create key value bears
            key = sheet.cell(1, col).value.lower().strip().replace(' ', '_')
            value = str(sheet.cell(row, col).value).lower().strip()

            item[key] = value

        # append rows
        list_items.append(item)

    for item in list_items:
        # remove unwanted keys
        del item['us']
        del item['serial_number']
        del item['المكان']

    models = []
    cpus = []
    rams = []
    hdds = []
    items = []
    for item in list_items:                
        handlingFields(models, 'labtop', item['manufacturer'], item['model'])

        if item['cpu'] != 'none':
            if item['cpu'].__contains__('intel'):
                val = item['cpu'].split('-')
                handlingFields(cpus, val[0], val[1])
            else:
                val = item['cpu'].split(' ')
                handlingFields(cpus, val[0], val[1])

        handlingFields(rams, item['ram'], item['ram-type'])
        handlingFields(hdds, item['hdd'], item['hdd1_type'])
        handlingFields(items, 'labtop', item['manufacturer'], item['model'], item['cpu'], 
                       item['ram'], item['ram-type'], item['hdd'], item['hdd1_type'], 
                       item['gpu'], item['touch_screen'], item['rotation_360deg'], 
                       item['illuminated_keyboard'], item['original_windows'], item['screen_resolution'], 
                       item['screen_size'], item['sound_type'], item['السعر'], item['خصم'], item['note'], True)

    print(items)
    return models, cpus, rams, hdds, items, list_items




from items.models import *



# models, cpus, rams, hdds, items, list_items = proccess_workbook('op.xlsx', 'الشيت القديم')

## generate items
# for i in list_items:
#     # models
#     typee, created = Types.objects.get_or_create(name='labtop')
#     manufacturer, created = Manufacturers.objects.get_or_create(name=i['manufacturer'])
#     model, created = Models.objects.get_or_create(typee=typee, manufacturer=manufacturer, name=i['model'])

#     # cpus
#     cpu = None
#     gen = None
#     if i['cpu'].__contains__('intel'):
#         val = i['cpu'].split('-')
#         cpu, created = CPUs.objects.get_or_create(name=val[0])
#         gen, created = CPUGenerations.objects.get_or_create(cpu=cpu, generation=val[1])
#     else:
#         val = i['cpu'].split(' ')
#         if val[0] != 'none':
#             cpu, created = CPUs.objects.get_or_create(name=val[0])
#             gen, created = CPUGenerations.objects.get_or_create(cpu=cpu, generation=' '.join(val[1:]))

#     # rams
#     if i['ram-type'] != 'none':
#         ram, created = Rams.objects.get_or_create(type=i['ram-type'], size=i['ram'])

#     # hdds
#     hdd = None
#     if i['hdd1_type'] != 'none':
#         hdd, created = HDDs.objects.get_or_create(type=i['hdd1_type'], size=i['hdd'])

#     resolution_type, created = ScreenResolution.objects.get_or_create(resolution_type=i['screen_resolution'])
#     sound_type, created = SoundTypes.objects.get_or_create(sound_type=i['sound_type'])

    
#     gpu, created = GPUs.objects.get_or_create(gpu_type=i['gpu'])

#     try:
#         item, created = Items.objects.get_or_create(model=model, cpu=gen, 
#                                         touch_screen=True if ['touch_screen'] in ['y', 'yes'] else False, 
#                                         rotation=i['rotation_360deg'], 
#                                         illuminated_keyboard=True if i['illuminated_keyboard'] in ['y', 'yes'] else False, 
#                                         original_windows=True if i['original_windows'] in ['y', 'yes'] else False, 
#                                         screen_resolution=resolution_type, screen_size=i['screen_size'], 
#                                         sound_type=sound_type, price=i['السعر'], disc=i['خصم'] if i['خصم'] != 'none' else 0, 
#                                         note=i['note'], is_available=True)
        
#         item.hdd.add(hdd)
#         item.ram.add(ram)
#         item.gpu.add(gpu)
#     except:
#         pass